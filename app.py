import streamlit as st
import pandas as pd
import numpy as np
import re
import requests
import unicodedata
import zipfile
from io import BytesIO
from openpyxl import load_workbook

# === Load default data from GitHub ===
@st.cache_data
def load_default_data():
    urls = {
        'file2': "https://raw.githubusercontent.com/phamtai1211/Thau_3PPharma/main/file2.xlsx",
        'file3': "https://raw.githubusercontent.com/phamtai1211/Thau_3PPharma/main/file3.xlsx",
        'file4': "https://raw.githubusercontent.com/phamtai1211/Thau_3PPharma/main/nhom_dieu_tri.xlsx"
    }
    data = {}
    for key, url in urls.items():
        resp = requests.get(url)
        resp.raise_for_status()
        data[key] = pd.read_excel(BytesIO(resp.content), engine='openpyxl')
    return data['file2'], data['file3'], data['file4']

file2, file3, file4 = load_default_data()

# === Text normalization helpers ===
def remove_diacritics(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def normalize_text(s: str) -> str:
    s = str(s)
    s = remove_diacritics(s).lower()
    return re.sub(r'\s+', '', s)

def normalize_active(name: str) -> str:
    return re.sub(r'\s+', ' ', re.sub(r'\(.*?\)', '', str(name))).strip().lower()

def normalize_concentration(conc: str) -> str:
    s = str(conc).lower().replace(',', '.')
    parts = [p.strip() for p in re.split(r'[;,]', s) if p.strip()]
    parts = [p for p in parts if re.search(r'\d', p)]
    if len(parts) >= 2 and re.search(r'(mg|mcg|g|%)', parts[0]) and 'ml' in parts[-1]:
        return parts[0].replace(' ', '') + '/' + parts[-1].replace(' ', '')
    return ''.join(p.replace(' ', '') for p in parts)

def normalize_group(grp: str) -> str:
    return re.sub(r'\D', '', str(grp)).strip()

# === Process uploaded Excel files ===
def process_uploaded(uploaded, df3_temp):
    # Determine sheet with most columns
    xls = pd.ExcelFile(uploaded, engine='openpyxl')
    sheet = max(xls.sheet_names, key=lambda s: pd.read_excel(uploaded, sheet_name=s, nrows=5, header=None, engine='openpyxl').shape[1])
    try:
        raw = pd.read_excel(uploaded, sheet_name=sheet, header=None, engine='openpyxl')
    except Exception:
        # Strip problematic style/theme files and dataValidations
        uploaded.seek(0)
        buf = BytesIO(uploaded.read())
        zf = zipfile.ZipFile(buf, 'r')
        out = BytesIO()
        with zipfile.ZipFile(out, 'w') as w:
            for item in zf.infolist():
                if item.filename.startswith('xl/styles') or item.filename.startswith('xl/theme'):
                    continue
                data = zf.read(item.filename)
                if item.filename.startswith('xl/worksheets/'):
                    data = re.sub(b'<dataValidations.*?</dataValidations>', b'', data, flags=re.DOTALL)
                w.writestr(item.filename, data)
        out.seek(0)
        wb = load_workbook(out, read_only=True, data_only=True)
        ws = wb[sheet]
        raw = pd.DataFrame(list(ws.iter_rows(values_only=True)))

    # Auto-detect header row among first 10
    header_idx = None
    scores = []
    for i in range(min(10, len(raw))):
        text = normalize_text(' '.join(raw.iloc[i].fillna('').astype(str).tolist()))
        sc = sum(kw in text for kw in ['tenhoatchat','soluong','nhomthuoc','nongdo'])
        scores.append((i, sc))
        if 'tenhoatchat' in text and 'soluong' in text:
            header_idx = i
            break
    if header_idx is None:
        idx, sc = max(scores, key=lambda x: x[1])
        header_idx = idx if sc > 0 else 0

    # Set header and body without user preview
    header = raw.iloc[header_idx].fillna('').astype(str).tolist()
    df_body = raw.iloc[header_idx+1:].copy()
    df_body.columns = header
    df_body = df_body.dropna(subset=header, how='all')
    df_body['_orig_idx'] = df_body.index
    df_body.reset_index(drop=True, inplace=True)

    # Map columns to standard names
    col_map = {}
    for c in df_body.columns:
        n = normalize_text(c)
        if 'tenhoatchat' in n or 'tenthanhphan' in n:
            col_map[c] = 'TÃªn hoáº¡t cháº¥t'
        elif 'nongdo' in n or 'hamluong' in n:
            col_map[c] = 'Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng'
        elif 'nhom' in n and 'thuoc' in n:
            col_map[c] = 'NhÃ³m thuá»‘c'
        elif 'soluong' in n:
            col_map[c] = 'Sá»‘ lÆ°á»£ng'
        elif 'duongdung' in n or 'duong' in n:
            col_map[c] = 'ÄÆ°á»ng dÃ¹ng'
        elif 'gia' in n:
            col_map[c] = 'GiÃ¡ káº¿ hoáº¡ch'
    df_body.rename(columns=col_map, inplace=True)

    # Normalize reference file2
    df2 = file2.copy()
    col_map2 = {}
    for c in df2.columns:
        n = normalize_text(c)
        if 'tenhoatchat' in n:
            col_map2[c] = 'TÃªn hoáº¡t cháº¥t'
        elif 'nongdo' in n or 'hamluong' in n:
            col_map2[c] = 'Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng'
        elif 'nhom' in n and 'thuoc' in n:
            col_map2[c] = 'NhÃ³m thuá»‘c'
        elif 'tensanpham' in n:
            col_map2[c] = 'TÃªn sáº£n pháº©m'
    df2.rename(columns=col_map2, inplace=True)

    # Add normalized merge keys
    for df_ in (df_body, df2):
        df_['active_norm'] = df_['TÃªn hoáº¡t cháº¥t'].apply(normalize_active)
        df_['conc_norm'] = df_['Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng'].apply(normalize_concentration)
        df_['group_norm'] = df_['NhÃ³m thuá»‘c'].apply(normalize_group)

    # Merge and deduplicate
    merged = pd.merge(df_body, df2, on=['active_norm','conc_norm','group_norm'], how='left', indicator=True)
    merged.drop_duplicates(subset=['_orig_idx'], keep='first', inplace=True)
    hosp = df3_temp[['TÃªn sáº£n pháº©m','Äá»‹a bÃ n','TÃªn KhÃ¡ch hÃ ng phá»¥ trÃ¡ch triá»ƒn khai']]
    merged = pd.merge(merged, hosp, on='TÃªn sáº£n pháº©m', how='left')

    # Prepare display and export DataFrames
    export_df = merged.drop(columns=['active_norm','conc_norm','group_norm','_merge','_orig_idx'])
    display_df = merged[merged['_merge']=='both'].drop(columns=['active_norm','conc_norm','group_norm','_merge','_orig_idx'])
    return display_df, export_df

# === Main UI ===
st.sidebar.title("Chá»©c nÄƒng")
option = st.sidebar.radio("Chá»n chá»©c nÄƒng", [
    "Lá»c Danh Má»¥c Tháº§u",
    "PhÃ¢n TÃ­ch Danh Má»¥c Tháº§u",
    "PhÃ¢n TÃ­ch Danh Má»¥c TrÃºng Tháº§u",
    "Äá» Xuáº¥t HÆ°á»›ng Triá»ƒn Khai"
])
# 1. Lá»c Danh Má»¥c Tháº§u
if option == "Lá»c Danh Má»¥c Tháº§u":
    st.header("ğŸ“‚ Lá»c Danh Má»¥c Tháº§u")
    # Chá»n Miá»n
    regions = sorted(file3["Miá»n"].dropna().unique())
    selected_region = st.selectbox("Chá»n Miá»n", regions)
    sub_df = file3[file3["Miá»n"] == selected_region] if selected_region else file3.copy()
    # Chá»n VÃ¹ng (náº¿u cÃ³)
    areas = sorted(sub_df["VÃ¹ng"].dropna().unique())
    selected_area = None
    if areas:
        selected_area = st.selectbox("Chá»n VÃ¹ng", ["(Táº¥t cáº£)"] + areas)
        if selected_area and selected_area != "(Táº¥t cáº£)":
            sub_df = sub_df[sub_df["VÃ¹ng"] == selected_area]
    # Chá»n Tá»‰nh
    provinces = sorted(sub_df["Tá»‰nh"].dropna().unique())
    selected_prov = st.selectbox("Chá»n Tá»‰nh", provinces)
    sub_df = sub_df[sub_df["Tá»‰nh"] == selected_prov] if selected_prov else sub_df
    # Chá»n Bá»‡nh viá»‡n/SYT
    hospitals = sorted(sub_df["Bá»‡nh viá»‡n/SYT"].dropna().unique())
    selected_hospital = st.selectbox("Chá»n Bá»‡nh viá»‡n/Sá»Ÿ Y Táº¿", hospitals)
    # Upload file danh má»¥c má»i tháº§u
    uploaded_file = st.file_uploader("Táº£i lÃªn file Danh Má»¥c Má»i Tháº§u (.xlsx)", type=["xlsx"])
    if uploaded_file is not None and selected_hospital:
        # Äá»c Excel vÃ  xÃ¡c Ä‘á»‹nh sheet chá»©a dá»¯ liá»‡u chÃ­nh
        xls = pd.ExcelFile(uploaded_file)
        sheet_name = None
        max_cols = 0
        for name in xls.sheet_names:
            try:
                df_test = xls.parse(name, nrows=1, header=None)
                cols = df_test.shape[1]
            except Exception:
                cols = 0
            if cols > max_cols:
                max_cols = cols
                sheet_name = name
        if sheet_name is None:
            st.error("âŒ KhÃ´ng tÃ¬m tháº¥y sheet dá»¯ liá»‡u phÃ¹ há»£p trong file.")
        else:
            # Äá»c toÃ n bá»™ sheet (khÃ´ng Ä‘áº·t header) Ä‘á»ƒ tÃ¬m dÃ²ng tiÃªu Ä‘á»
            df_raw = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
            header_index = None
            for i in range(10):
                row = " ".join(df_raw.iloc[i].fillna('').astype(str).tolist())
                if "TÃªn hoáº¡t cháº¥t" in row and "Sá»‘ lÆ°á»£ng" in row:
                    header_index = i
                    break
            if header_index is None:
                st.error("âŒ KhÃ´ng xÃ¡c Ä‘á»‹nh Ä‘Æ°á»£c dÃ²ng tiÃªu Ä‘á» trong file.")
            else:
                # Táº¡o DataFrame vá»›i header chÃ­nh xÃ¡c
                header = df_raw.iloc[header_index].tolist()
                df_all = df_raw.iloc[header_index+1:].reset_index(drop=True)
                df_all.columns = header
                # Bá» cÃ¡c dÃ²ng trá»‘ng hoÃ n toÃ n (náº¿u cÃ³)
                df_all = df_all.dropna(how='all').reset_index(drop=True)
                # So sÃ¡nh 3 cá»™t (hoáº¡t cháº¥t, hÃ m lÆ°á»£ng, nhÃ³m thuá»‘c) vá»›i danh má»¥c cÃ´ng ty (file2)
                df_all["active_norm"] = df_all["TÃªn hoáº¡t cháº¥t"].apply(normalize_active)
                df_all["conc_norm"] = df_all["Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng"].apply(normalize_concentration)
                df_all["group_norm"] = df_all["NhÃ³m thuá»‘c"].apply(normalize_group)
                df_comp = file2.copy()
                df_comp["active_norm"] = df_comp["TÃªn hoáº¡t cháº¥t"].apply(normalize_active)
                df_comp["conc_norm"] = df_comp["Ná»“ng Ä‘á»™/HÃ m lÆ°á»£ng"].apply(normalize_concentration)
                df_comp["group_norm"] = df_comp["NhÃ³m thuá»‘c"].apply(normalize_group)
                # Inner merge Ä‘á»ƒ giá»¯ láº¡i cÃ¡c dÃ²ng khá»›p vá»›i danh má»¥c cÃ´ng ty
                merged_df = pd.merge(df_all, df_comp, on=["active_norm", "conc_norm", "group_norm"], how="inner", suffixes=(None, "_comp"))
                # Chá»n cÃ¡c cá»™t gá»‘c + tÃªn sáº£n pháº©m (brand), Ä‘á»“ng thá»i gáº¯n Äá»‹a bÃ n vÃ  KhÃ¡ch hÃ ng phá»¥ trÃ¡ch
                result_columns = df_all.columns.tolist() + ["TÃªn sáº£n pháº©m"]
                result_df = merged_df[result_columns].copy()
                # ThÃªm thÃ´ng tin Äá»‹a bÃ n, KhÃ¡ch hÃ ng phá»¥ trÃ¡ch tá»« file3
                hosp_data = file3[file3["Bá»‡nh viá»‡n/SYT"] == selected_hospital][["TÃªn sáº£n pháº©m", "Äá»‹a bÃ n", "TÃªn KhÃ¡ch hÃ ng phá»¥ trÃ¡ch triá»ƒn khai"]]
                result_df = pd.merge(result_df, hosp_data, on="TÃªn sáº£n pháº©m", how="left")
                # TÃ­nh cá»™t "Tá»· trá»ng SL/DM Tá»•ng"
                # Láº­p báº£ng tá»•ng sá»‘ lÆ°á»£ng theo NhÃ³m Ä‘iá»u trá»‹ cho toÃ n bá»™ danh má»¥c tháº§u (df_all)
                # Ãnh xáº¡ hoáº¡t cháº¥t -> nhÃ³m Ä‘iá»u trá»‹ tá»« file4
                treat_map = { normalize_active(a): grp for a, grp in zip(file4["Hoáº¡t cháº¥t"], file4["NhÃ³m Ä‘iá»u trá»‹"]) }
                group_total = {}
                for _, row in df_all.iterrows():
                    act = normalize_active(row["TÃªn hoáº¡t cháº¥t"])
                    group = treat_map.get(act)
                    qty = pd.to_numeric(row.get("Sá»‘ lÆ°á»£ng", 0), errors='coerce')
                    if pd.isna(qty):
                        qty = 0
                    if group:
                        group_total[group] = group_total.get(group, 0) + float(qty)
                # TÃ­nh tá»· trá»ng cho tá»«ng dÃ²ng káº¿t quáº£
                ratios = []
                for _, row in result_df.iterrows():
                    act = normalize_active(row["TÃªn hoáº¡t cháº¥t"])
                    group = treat_map.get(act)
                    qty = pd.to_numeric(row.get("Sá»‘ lÆ°á»£ng", 0), errors='coerce')
                    if pd.isna(qty) or group is None or group not in group_total or group_total[group] == 0:
                        ratios.append(None)
                    else:
                        ratio = float(qty) / group_total[group]
                        ratios.append(f"{ratio:.2%}")
                result_df["Tá»· trá»ng SL/DM Tá»•ng"] = ratios
                # Hiá»ƒn thá»‹ káº¿t quáº£ lá»c vÃ  nÃºt táº£i vá»
                st.success(f"âœ… ÄÃ£ lá»c Ä‘Æ°á»£c {len(result_df)} má»¥c thuá»‘c thuá»™c danh má»¥c cÃ´ng ty.")
                st.dataframe(result_df.head(10))
                # Xuáº¥t file Excel káº¿t quáº£
                output = BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    result_df.to_excel(writer, sheet_name="KetQuaLoc", index=False)
                st.download_button("â¬‡ï¸ Táº£i File Káº¿t Quáº£", data=output.getvalue(), file_name="Ketqua_loc.xlsx")
                # LÆ°u DataFrame Ä‘Ã£ lá»c vÃ o session_state Ä‘á»ƒ dÃ¹ng cho phÃ¢n tÃ­ch
                st.session_state["filtered_df"] = result_df
                st.session_state["selected_hospital"] = selected_hospital

# 2. PhÃ¢n TÃ­ch Danh Má»¥c Tháº§u
elif option == "PhÃ¢n TÃ­ch Danh Má»¥c Tháº§u":
    st.header("ğŸ“Š PhÃ¢n TÃ­ch Danh Má»¥c Tháº§u")
    if "filtered_df" not in st.session_state:
        st.info("Vui lÃ²ng thá»±c hiá»‡n bÆ°á»›c 'Lá»c Danh Má»¥c Tháº§u' trÆ°á»›c.")
    else:
        df_filtered = st.session_state["filtered_df"].copy()
        # Äáº£m báº£o kiá»ƒu dá»¯ liá»‡u sá»‘
        df_filtered["Sá»‘ lÆ°á»£ng"] = pd.to_numeric(df_filtered["Sá»‘ lÆ°á»£ng"], errors='coerce').fillna(0)
        df_filtered["GiÃ¡ káº¿ hoáº¡ch"] = pd.to_numeric(df_filtered["GiÃ¡ káº¿ hoáº¡ch"], errors='coerce').fillna(0)
        # ThÃªm cá»™t trá»‹ giÃ¡ = Sá»‘ lÆ°á»£ng * GiÃ¡ káº¿ hoáº¡ch
        df_filtered["Trá»‹ giÃ¡"] = df_filtered["Sá»‘ lÆ°á»£ng"] * df_filtered["GiÃ¡ káº¿ hoáº¡ch"]
        # Biá»ƒu Ä‘á»“ 1: NhÃ³m tháº§u sá»­ dá»¥ng nhiá»u nháº¥t theo trá»‹ giÃ¡
        group_val = df_filtered.groupby("NhÃ³m thuá»‘c")["Trá»‹ giÃ¡"].sum().reset_index().sort_values("Trá»‹ giÃ¡", ascending=False)
        fig1 = px.bar(group_val, x="NhÃ³m thuá»‘c", y="Trá»‹ giÃ¡", title="Trá»‹ giÃ¡ theo NhÃ³m tháº§u (gÃ³i tháº§u)")
        st.plotly_chart(fig1, use_container_width=True)
        # Biá»ƒu Ä‘á»“ 2: PhÃ¢n tÃ­ch Ä‘Æ°á»ng dÃ¹ng (tiÃªm/uá»‘ng) theo trá»‹ giÃ¡
        # XÃ¡c Ä‘á»‹nh loáº¡i Ä‘Æ°á»ng dÃ¹ng cho tá»«ng má»¥c (TiÃªm, Uá»‘ng hoáº·c KhÃ¡c)
        route_df = df_filtered.copy()
        def classify_route(route_str):
            route = str(route_str).lower()
            if "tiÃªm" in route:
                return "TiÃªm"
            elif "uá»‘ng" in route:
                return "Uá»‘ng"
            else:
                return "KhÃ¡c"
        route_df["Loáº¡i Ä‘Æ°á»ng dÃ¹ng"] = route_df["ÄÆ°á»ng dÃ¹ng"].apply(classify_route)
        route_val = route_df.groupby("Loáº¡i Ä‘Æ°á»ng dÃ¹ng")["Trá»‹ giÃ¡"].sum().reset_index()
        fig2 = px.pie(route_val, names="Loáº¡i Ä‘Æ°á»ng dÃ¹ng", values="Trá»‹ giÃ¡", title="Tá»· trá»ng trá»‹ giÃ¡ theo Ä‘Æ°á»ng dÃ¹ng")
        st.plotly_chart(fig2, use_container_width=True)
        # Biá»ƒu Ä‘á»“ 3: Top 10 hoáº¡t cháº¥t theo Sá»‘ lÆ°á»£ng
        top_active_qty = df_filtered.groupby("TÃªn hoáº¡t cháº¥t")["Sá»‘ lÆ°á»£ng"].sum().reset_index().sort_values("Sá»‘ lÆ°á»£ng", ascending=False).head(10)
        fig3 = px.bar(top_active_qty, x="TÃªn hoáº¡t cháº¥t", y="Sá»‘ lÆ°á»£ng", title="Top 10 Hoáº¡t cháº¥t (theo Sá»‘ lÆ°á»£ng)")
        st.plotly_chart(fig3, use_container_width=True)
        # Biá»ƒu Ä‘á»“ 4: Top 10 hoáº¡t cháº¥t theo Trá»‹ giÃ¡
        top_active_val = df_filtered.groupby("TÃªn hoáº¡t cháº¥t")["Trá»‹ giÃ¡"].sum().reset_index().sort_values("Trá»‹ giÃ¡", ascending=False).head(10)
        fig4 = px.bar(top_active_val, x="TÃªn hoáº¡t cháº¥t", y="Trá»‹ giÃ¡", title="Top 10 Hoáº¡t cháº¥t (theo Trá»‹ giÃ¡)")
        st.plotly_chart(fig4, use_container_width=True)
        # Biá»ƒu Ä‘á»“ 5: PhÃ¢n tÃ­ch NhÃ³m Ä‘iá»u trá»‹ vÃ  top 10 sáº£n pháº©m
        # Gáº¯n cá»™t NhÃ³m Ä‘iá»u trá»‹ cho tá»«ng má»¥c
        treat_map = { normalize_active(a): grp for a, grp in zip(file4["Hoáº¡t cháº¥t"], file4["NhÃ³m Ä‘iá»u trá»‹"]) }
        df_filtered["NhÃ³m Ä‘iá»u trá»‹"] = df_filtered["TÃªn hoáº¡t cháº¥t"].apply(lambda x: treat_map.get(normalize_active(x), "KhÃ¡c"))
        # Tá»•ng trá»‹ giÃ¡ theo nhÃ³m Ä‘iá»u trá»‹
        treat_val = df_filtered.groupby("NhÃ³m Ä‘iá»u trá»‹")["Trá»‹ giÃ¡"].sum().reset_index().sort_values("Trá»‹ giÃ¡", ascending=False)
        fig5 = px.bar(treat_val, x="Trá»‹ giÃ¡", y="NhÃ³m Ä‘iá»u trá»‹", orientation='h', title="Trá»‹ giÃ¡ theo NhÃ³m Ä‘iá»u trá»‹")
        st.plotly_chart(fig5, use_container_width=True)
        # Chá»n nhÃ³m Ä‘iá»u trá»‹ Ä‘á»ƒ xem Top 10 sáº£n pháº©m
        groups = treat_val["NhÃ³m Ä‘iá»u trá»‹"].tolist()
        selected_grp = st.selectbox("Chá»n NhÃ³m Ä‘iá»u trá»‹ Ä‘á»ƒ xem Top 10 sáº£n pháº©m", groups)
        if selected_grp:
            top_products = df_filtered[df_filtered["NhÃ³m Ä‘iá»u trá»‹"] == selected_grp].groupby("TÃªn sáº£n pháº©m")["Trá»‹ giÃ¡"].sum().reset_index().sort_values("Trá»‹ giÃ¡", ascending=False).head(10)
            fig6 = px.bar(top_products, x="Trá»‹ giÃ¡", y="TÃªn sáº£n pháº©m", orientation='h', title=f"Top 10 sáº£n pháº©m - NhÃ³m {selected_grp}")
            st.plotly_chart(fig6, use_container_width=True)
        # Biá»ƒu Ä‘á»“ 6: Hiá»‡u quáº£ theo TÃªn khÃ¡ch hÃ ng phá»¥ trÃ¡ch triá»ƒn khai (tá»•ng trá»‹ giÃ¡ theo ngÆ°á»i phá»¥ trÃ¡ch)
        rep_val = df_filtered.groupby("TÃªn KhÃ¡ch hÃ ng phá»¥ trÃ¡ch triá»ƒn khai")["Trá»‹ giÃ¡"].sum().reset_index().sort_values("Trá»‹ giÃ¡", ascending=False)
        fig7 = px.bar(rep_val, x="Trá»‹ giÃ¡", y="TÃªn KhÃ¡ch hÃ ng phá»¥ trÃ¡ch triá»ƒn khai", orientation='h', title="Trá»‹ giÃ¡ theo KhÃ¡ch hÃ ng phá»¥ trÃ¡ch")
        st.plotly_chart(fig7, use_container_width=True)

# 3. PhÃ¢n TÃ­ch Danh Má»¥c TrÃºng Tháº§u
elif option == "PhÃ¢n TÃ­ch Danh Má»¥c TrÃºng Tháº§u":
    st.header("ğŸ† PhÃ¢n TÃ­ch Danh Má»¥c TrÃºng Tháº§u")
    win_file = st.file_uploader("Táº£i lÃªn file Káº¿t Quáº£ TrÃºng Tháº§u (.xlsx)", type=["xlsx"])
    invite_file = st.file_uploader("Táº£i lÃªn file Danh Má»¥c Má»i Tháº§u (Ä‘á»ƒ Ä‘á»‘i chiáº¿u, tÃ¹y chá»n)", type=["xlsx"])
    if win_file is not None:
        # XÃ¡c Ä‘á»‹nh sheet chÃ­nh cá»§a file trÃºng tháº§u
        xls_win = pd.ExcelFile(win_file)
        win_sheet = xls_win.sheet_names[0]
        max_cols = 0
        for name in xls_win.sheet_names:
            try:
                df_test = xls_win.parse(name, nrows=1, header=None)
                cols = df_test.shape[1]
            except:
                cols = 0
            if cols > max_cols:
                max_cols = cols
                win_sheet = name
        # Äá»c toÃ n bá»™ sheet vÃ  xÃ¡c Ä‘á»‹nh dÃ²ng tiÃªu Ä‘á»
        df_win_raw = pd.read_excel(win_file, sheet_name=win_sheet, header=None)
        header_idx = None
        for i in range(10):
            row_text = " ".join(df_win_raw.iloc[i].fillna('').astype(str).tolist())
            if "TÃªn hoáº¡t cháº¥t" in row_text and "NhÃ  tháº§u trÃºng" in row_text:
                header_idx = i
                break
        if header_idx is None:
            st.error("âŒ KhÃ´ng xÃ¡c Ä‘á»‹nh Ä‘Æ°á»£c tiÃªu Ä‘á» cá»™t trong file trÃºng tháº§u.")
        else:
            header = df_win_raw.iloc[header_idx].tolist()
            df_win = df_win_raw.iloc[header_idx+1:].reset_index(drop=True)
            df_win.columns = header
            df_win = df_win.dropna(how='all').reset_index(drop=True)
            # Chuyá»ƒn kiá»ƒu sá»‘ cho Sá»‘ lÆ°á»£ng vÃ  giÃ¡
            df_win["Sá»‘ lÆ°á»£ng"] = pd.to_numeric(df_win.get("Sá»‘ lÆ°á»£ng", 0), errors='coerce').fillna(0)
            # XÃ¡c Ä‘á»‹nh cá»™t giÃ¡ trÃºng (náº¿u khÃ´ng cÃ³ thÃ¬ dÃ¹ng GiÃ¡ káº¿ hoáº¡ch)
            price_col = None
            for col in df_win.columns:
                if "GiÃ¡ trÃºng" in str(col):
                    price_col = col
                    break
            if price_col is None:
                price_col = "GiÃ¡ káº¿ hoáº¡ch"
            df_win[price_col] = pd.to_numeric(df_win.get(price_col, 0), errors='coerce').fillna(0)
            # TÃ­nh trá»‹ giÃ¡ trÃºng tháº§u má»—i má»¥c
            df_win["Trá»‹ giÃ¡"] = df_win["Sá»‘ lÆ°á»£ng"] * df_win[price_col]
            # Biá»ƒu Ä‘á»“: Top 20 nhÃ  tháº§u trÃºng trá»‹ giÃ¡ cao nháº¥t
            win_val = df_win.groupby("NhÃ  tháº§u trÃºng")["Trá»‹ giÃ¡"].sum().reset_index().sort_values("Trá»‹ giÃ¡", ascending=False).head(20)
            fig_w1 = px.bar(win_val, x="Trá»‹ giÃ¡", y="NhÃ  tháº§u trÃºng", orientation='h', title="Top 20 NhÃ  tháº§u trÃºng (theo trá»‹ giÃ¡)")
            st.plotly_chart(fig_w1, use_container_width=True)
            # Biá»ƒu Ä‘á»“: PhÃ¢n tÃ­ch theo nhÃ³m Ä‘iá»u trá»‹ (cÆ¡ cáº¥u trá»‹ giÃ¡)
            df_win["NhÃ³m Ä‘iá»u trá»‹"] = df_win["TÃªn hoáº¡t cháº¥t"].apply(lambda x: treat_map.get(normalize_active(x), "KhÃ¡c"))
            treat_win = df_win.groupby("NhÃ³m Ä‘iá»u trá»‹")["Trá»‹ giÃ¡"].sum().reset_index().sort_values("Trá»‹ giÃ¡", ascending=False)
            fig_w2 = px.pie(treat_win, names="NhÃ³m Ä‘iá»u trá»‹", values="Trá»‹ giÃ¡", title="CÆ¡ cáº¥u trá»‹ giÃ¡ theo NhÃ³m Ä‘iá»u trá»‹ (TrÃºng tháº§u)")
            st.plotly_chart(fig_w2, use_container_width=True)
            # Náº¿u cÃ³ upload danh má»¥c má»i tháº§u Ä‘á»ƒ Ä‘á»‘i chiáº¿u
            if invite_file is not None:
                xls_inv = pd.ExcelFile(invite_file)
                inv_sheet = xls_inv.sheet_names[0]
                df_inv_raw = pd.read_excel(invite_file, sheet_name=inv_sheet, header=None)
                header_idx2 = None
                for i in range(10):
                    row_text = " ".join(df_inv_raw.iloc[i].fillna('').astype(str).tolist())
                    if "TÃªn hoáº¡t cháº¥t" in row_text and "Sá»‘ lÆ°á»£ng" in row_text:
                        header_idx2 = i
                        break
                if header_idx2 is not None:
                    header2 = df_inv_raw.iloc[header_idx2].tolist()
                    df_inv_full = df_inv_raw.iloc[header_idx2+1:].reset_index(drop=True)
                    df_inv_full.columns = header2
                    df_inv_full = df_inv_full.dropna(how='all').reset_index(drop=True)
                    # So sÃ¡nh cÃ¡c má»¥c khÃ´ng trÃºng (cÃ³ trong má»i tháº§u nhÆ°ng khÃ´ng cÃ³ trong trÃºng tháº§u)
                    if "MÃ£ pháº§n (LÃ´)" in df_inv_full.columns and "MÃ£ pháº§n (LÃ´)" in df_win.columns:
                        inv_ids = set(df_inv_full["MÃ£ pháº§n (LÃ´)"].astype(str))
                        win_ids = set(df_win["MÃ£ pháº§n (LÃ´)"].astype(str))
                        missing_ids = inv_ids - win_ids
                        missing_items = df_inv_full[df_inv_full["MÃ£ pháº§n (LÃ´)"].astype(str).isin(missing_ids)]
                    else:
                        # DÃ¹ng káº¿t há»£p hoáº¡t cháº¥t + hÃ m lÆ°á»£ng Ä‘á»ƒ Ä‘á»‘i chiáº¿u náº¿u khÃ´ng cÃ³ MÃ£ pháº§n
                        inv_keys = df_inv_full["TÃªn hoáº¡t cháº¥t"].astype(str) + df_inv_full["Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng"].astype(str)
                        win_keys = df_win["TÃªn hoáº¡t cháº¥t"].astype(str) + df_win["Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng"].astype(str)
                        missing_mask = ~inv_keys.isin(win_keys)
                        missing_items = df_inv_full[missing_mask]
                    if not missing_items.empty:
                        st.write("**CÃ¡c thuá»‘c má»i tháº§u khÃ´ng cÃ³ nhÃ  tháº§u trÃºng:**")
                        st.dataframe(missing_items[["TÃªn hoáº¡t cháº¥t", "Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng", "Sá»‘ lÆ°á»£ng", "GiÃ¡ káº¿ hoáº¡ch"]])
                        st.write(f"ğŸ“Œ Sá»‘ lÆ°á»£ng thuá»‘c khÃ´ng trÃºng tháº§u: {len(missing_items)}")
                        # LÆ°u vÃ o session_state Ä‘á»ƒ dÃ¹ng cho Ä‘á» xuáº¥t
                        st.session_state["missing_items"] = missing_items
                    else:
                        st.write("âœ… Táº¥t cáº£ thuá»‘c má»i tháº§u Ä‘á»u Ä‘Ã£ cÃ³ nhÃ  tháº§u trÃºng.")

# 4. Äá» Xuáº¥t HÆ°á»›ng Triá»ƒn Khai
elif option == "Äá» Xuáº¥t HÆ°á»›ng Triá»ƒn Khai":
    st.header("ğŸ’¡ Äá» Xuáº¥t HÆ°á»›ng Triá»ƒn Khai")
    if "filtered_df" not in st.session_state:
        st.info("Vui lÃ²ng thá»±c hiá»‡n phÃ¢n tÃ­ch trÆ°á»›c Ä‘á»ƒ cÃ³ dá»¯ liá»‡u.")
    else:
        df_filtered = st.session_state["filtered_df"]
        hospital = st.session_state.get("selected_hospital", "")
        # Danh sÃ¡ch Ä‘á» xuáº¥t
        suggestions_yes = []  # nÃªn triá»ƒn khai
        suggestions_no = []   # khÃ´ng nÃªn triá»ƒn khai
        # 1. CÃ¡c sáº£n pháº©m trong danh má»¥c cÃ´ng ty táº¡i bá»‡nh viá»‡n nhÆ°ng chÆ°a cÃ³ trong danh má»¥c má»i tháº§u
        hosp_products = set(file3[file3["Bá»‡nh viá»‡n/SYT"] == hospital]["TÃªn sáº£n pháº©m"])
        included_products = set(df_filtered["TÃªn sáº£n pháº©m"])
        not_included = hosp_products - included_products
        # XÃ¡c Ä‘á»‹nh nhÃ³m bá»‡nh viá»‡n tÆ°Æ¡ng tá»± (cÃ¹ng Miá»n, cÃ¹ng loáº¡i SYT hoáº·c BV)
        hosp_info = file3[file3["Bá»‡nh viá»‡n/SYT"] == hospital].iloc[0] if not file3[file3["Bá»‡nh viá»‡n/SYT"] == hospital].empty else None
        similar_df = file3.copy()
        if hosp_info is not None:
            if "SYT" in hospital:
                # cÃ¡c Sá»Ÿ Y Táº¿ khÃ¡c trong cÃ¹ng Miá»n
                similar_df = similar_df[similar_df["Bá»‡nh viá»‡n/SYT"].str.contains("SYT") & (similar_df["Miá»n"] == hosp_info["Miá»n"])]
            else:
                # cÃ¡c Bá»‡nh viá»‡n khÃ¡c (khÃ´ng pháº£i SYT) trong cÃ¹ng Miá»n
                similar_df = similar_df[~similar_df["Bá»‡nh viá»‡n/SYT"].str.contains("SYT") & (similar_df["Miá»n"] == hosp_info["Miá»n"])]
        for prod in not_included:
            if prod in set(similar_df["TÃªn sáº£n pháº©m"]):
                suggestions_yes.append(f"- NÃªn triá»ƒn khai **{prod}**: Sáº£n pháº©m chÆ°a cÃ³ trong tháº§u cá»§a {hospital}, nhÆ°ng nhiá»u Ä‘Æ¡n vá»‹ tÆ°Æ¡ng tá»± Ä‘Ã£ cÃ³ nhu cáº§u.")
            else:
                suggestions_no.append(f"- ChÆ°a cáº§n triá»ƒn khai **{prod}**: Sáº£n pháº©m chÆ°a cÃ³ trong tháº§u {hospital} vÃ  chÆ°a phá»• biáº¿n á»Ÿ nhÃ³m bá»‡nh viá»‡n tÆ°Æ¡ng tá»±.")
        # 2. CÃ¡c sáº£n pháº©m má»i tháº§u nhÆ°ng khÃ´ng cÃ³ nhÃ  tháº§u trÃºng (náº¿u cÃ³)
        if "missing_items" in st.session_state:
            missing_items = st.session_state["missing_items"]
            for _, row in missing_items.iterrows():
                suggestions_yes.append(f"- Thá»­ triá»ƒn khai **{row['TÃªn hoáº¡t cháº¥t']}**: Thuá»‘c Ä‘Æ°á»£c má»i tháº§u {hospital} nhÆ°ng chÆ°a cÃ³ nhÃ  tháº§u trÃºng, cÃ³ thá»ƒ lÃ  cÆ¡ há»™i Ä‘Æ°a sáº£n pháº©m vÃ o.")
        # 3. CÃ¡c sáº£n pháº©m cÃ³ Ä‘á»‘i thá»§ trÃºng tháº§u (cÃ´ng ty chÆ°a trÃºng)
        # Giáº£ sá»­ cÃ´ng ty theo dÃµi cÃ¡c sáº£n pháº©m Ä‘Ã£ Ä‘Æ°á»£c Ä‘Æ°a vÃ o tháº§u (df_filtered), náº¿u khÃ´ng trÃºng tháº§u cÃ³ thá»ƒ cÃ¢n nháº¯c má»©c Ä‘á»™ Æ°u tiÃªn
        if "missing_items" in st.session_state or "filtered_df" in st.session_state:
            # Náº¿u má»™t sáº£n pháº©m cÃ³ máº·t trong danh má»¥c má»i tháº§u (cá»§a cÃ´ng ty) nhÆ°ng cÃ´ng ty khÃ´ng trÃºng -> Ä‘á»‘i thá»§ Ä‘Ã£ trÃºng
            # (ÄÆ¡n giáº£n coi nhÆ° má»i má»¥c trong df_filtered lÃ  cÃ´ng ty cÃ³ tham gia, náº¿u khÃ´ng náº±m trong missing_items tá»©c lÃ  cÃ³ ngÆ°á»i trÃºng)
            if "missing_items" in st.session_state:
                lost_df = df_filtered.copy()
                for _, miss in st.session_state["missing_items"].iterrows():
                    # loáº¡i cÃ¡c má»¥c khÃ´ng ai trÃºng (Ä‘Ã£ xá»­ lÃ½ á»Ÿ trÃªn)
                    lost_df = lost_df[~((lost_df["TÃªn hoáº¡t cháº¥t"] == miss["TÃªn hoáº¡t cháº¥t"]) & (lost_df["Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng"] == miss["Ná»“ng Ä‘á»™/hÃ m lÆ°á»£ng"]))]
            else:
                lost_df = df_filtered
            # Táº¥t cáº£ má»¥c cÃ²n láº¡i trong lost_df coi nhÆ° cÃ³ Ä‘á»‘i thá»§ trÃºng
            for _, row in lost_df.iterrows():
                suggestions_no.append(f"- Háº¡n cháº¿ táº­p trung **{row['TÃªn hoáº¡t cháº¥t']}**: ÄÃ£ cÃ³ Ä‘á»‘i thá»§ trÃºng tháº§u táº¡i {hospital}, cáº§n cÃ¢n nháº¯c náº¿u khÃ´ng cÃ³ lá»£i tháº¿ cáº¡nh tranh.")
        # Hiá»ƒn thá»‹ Ä‘á» xuáº¥t
        st.subheader("ğŸ”¸ Äá» xuáº¥t nÃªn triá»ƒn khai")
        if suggestions_yes:
            st.markdown("\n".join(suggestions_yes))
        else:
            st.write("KhÃ´ng cÃ³ sáº£n pháº©m má»›i nÃ o cáº§n triá»ƒn khai thÃªm táº¡i thá»i Ä‘iá»ƒm nÃ y.")
        st.subheader("ğŸ”¹ Äá» xuáº¥t khÃ´ng nÃªn triá»ƒn khai")
        if suggestions_no:
            st.markdown("\n".join(suggestions_no))
        else:
            st.write("KhÃ´ng cÃ³ sáº£n pháº©m nÃ o cáº§n ngá»«ng triá»ƒn khai; tiáº¿p tá»¥c duy trÃ¬ cÃ¡c danh má»¥c hiá»‡n cÃ³.")
